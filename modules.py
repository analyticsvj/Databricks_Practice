# readery.py
# import os
# import logging

# logger = logging.getLogger("PETFileLoader")

# def get_dbutils():
#     try:
#         import IPython
#         return IPython.get_ipython().user_ns["dbutils"]
#     except Exception:
#         try:
#             from dbruntime.dbutils import DBUtils
#             return DBUtils(spark)
#         except ImportError:
#             raise RuntimeError("dbutils is not available in this context.")

# def download_to_tmp(source_path: str) -> str:
#     """
#     Downloads a file from S3 or DBFS to local /dbfs/tmp using dbutils.fs.cp.
#     """
#     dbutils = get_dbutils()  # 🔁 Safe call here

#     filename = os.path.basename(source_path)
#     dbfs_path = f"dbfs:/tmp/{filename}"
#     logger.info(f"Copying {source_path} → {dbfs_path}")
#     dbutils.fs.cp(source_path, dbfs_path, recurse=False)
#     return f"/dbfs/tmp/{filename}"
import os
import logging
from IPython import get_ipython

logger = logging.getLogger("PETFileLoader")

# ✅ Get dbutils from notebook context
try:
    dbutils = get_ipython().user_ns["dbutils"]
except Exception:
    raise RuntimeError("dbutils is not available. Ensure you're running inside a Databricks notebook.")

# ✅ Function to copy from S3 or DBFS to /dbfs/tmp
def download_to_tmp(source_path: str) -> str:
    """
    Downloads a file from S3 or DBFS to local /dbfs/tmp using dbutils.fs.cp.
    Returns local path e.g. /dbfs/tmp/myfile.xlsx
    """
    filename = os.path.basename(source_path)
    dbfs_path = f"dbfs:/tmp/{filename}"
    logger.info(f"Copying {source_path} → {dbfs_path}")
    dbutils.fs.cp(source_path, dbfs_path, recurse=False)
    return f"/dbfs/tmp/{filename}"
=============================================================================
# validators.py
================================
from typing import List, Dict, Optional
from datetime import datetime
import logging
from pyspark.sql import SparkSession
from pyspark.sql.functions import col

logger = logging.getLogger("PETFileLoader")

# def validate_required_sheets(sheetnames: List[str]) -> bool:
#     required_sheets = {"Specification", "Input Fields", "DPS-CDP Params"}
#     missing = required_sheets - set(sheetnames)
#     if missing:
#         raise ValueError(f"Missing required sheets: {missing}")
#     return True

def validate_required_sheets(sheetnames: list) -> bool:
    required_sheets = {"Specification", "Input Fields", "DPS-CDP Params"}
    missing = required_sheets - set(sheetnames)
    if missing:
        raise ValueError(f"Missing required sheets: {missing}")
    return True

def validate_specification_format(spec_sheet) -> Dict[str, str]:
    expected_keys = [
        "Source Tenant", "Dataset Name", "Data Sharing Agreement (DSA)", "Purpose",
        "Privacy Domain - Encryption Key", "Destination Tenant", "ExternalID",
        "Allow Missing Table", "Priority", "Policy"
    ]
    spec_rows = list(spec_sheet.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True))
    if len(spec_rows) < 10:
        raise ValueError(f"Specification sheet must contain 10 rows, found {len(spec_rows)}.")
    output = {}
    for i, expected_key in enumerate(expected_keys):
        actual_key, actual_value = spec_rows[i]
        if actual_key != expected_key:
            raise ValueError(f"Row {i+1} mismatch: expected '{expected_key}', found '{actual_key}'")
        if not actual_value:
            raise ValueError(f"Missing value for '{expected_key}' at row {i+1}")
        output[actual_key] = str(actual_value).strip()
    return output

def check_database_exists(spark: SparkSession, db_name: str) -> bool:
    return db_name.lower() in [row.databaseName.lower() for row in spark.sql("SHOW DATABASES").collect()]

def check_table_and_column_exist(spark: SparkSession, db_name: str, table_name: str, column_name: str) -> bool:
    try:
        if not spark.sql(f"SHOW TABLES IN {db_name}").filter(col("tableName") == table_name).count():
            return False
        df_schema = spark.table(f"{db_name}.{table_name}").schema
        return any(
            f.name == column_name or
            ("." in column_name and column_name.startswith(f.name + "."))
            for f in df_schema.fields
        )
    except Exception as e:
        logger.exception(f"Table/Column check failed: {e}")
        raise

def validate_incremental_timestamp_field(spark: SparkSession, db_name: str, field: Optional[str]) -> None:
    if not field:
        raise ValueError("Incremental Timestamp Field is missing")
    parts = field.split(".")
    if len(parts) == 2:
        table_name, column_name = parts
        if not check_table_and_column_exist(spark, db_name, table_name, column_name):
            raise ValueError(f"Incremental Timestamp Field not found: {field}")
    else:
        raise ValueError(f"Invalid format for Incremental Timestamp Field: {field}")

def get_last_process_map(spark: SparkSession, dataset_id: str, start_process_from_date_str: Optional[str]) -> Optional[Dict[str, datetime]]:
    try:
        df_existing = spark.sql(
            f"SELECT last_process_datetime FROM pet.ctrl_dataset_config_vj WHERE dataset_id = '{dataset_id}'"
        )
        if df_existing.count() > 0:
            current_map = df_existing.collect()[0]["last_process_datetime"]
            if current_map and ("initialise" in current_map or dataset_id.lower() in current_map):
                return current_map

        if not start_process_from_date_str:
            return None

        try:
            start_process_from_date = datetime.strptime(start_process_from_date_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            start_process_from_date = datetime.strptime(start_process_from_date_str, "%Y-%m-%d")

        return {"initialise": start_process_from_date}
    except Exception as e:
        logger.exception(f"Error checking last_process_datetime: {e}")
        raise
========================================================================
# writers.py
==============================

from pyspark.sql import SparkSession, DataFrame
from pyspark.sql.types import StructType
import logging

logger = logging.getLogger("PETWriter")

def write_config_table(spark: SparkSession, df: DataFrame, schema: StructType, overwrite: bool) -> None:
    """Write configuration data to the Delta table."""
    try:
        df_to_write = spark.createDataFrame(df.rdd, schema=schema)
        df_to_write.write.format("delta").mode("overwrite" if overwrite else "append").saveAsTable("pet.ctrl_dataset_config_vj")
        logger.info("Configuration data written successfully.")
    except Exception as e:
        logger.exception(f"Failed to write to pet.ctrl_dataset_config_vj: {e}")
        raise

def write_input_fields_table(spark: SparkSession, df: DataFrame, dataset_id: str) -> None:
    """Write input fields to the Delta table."""
    try:
        df = df.copy()
        df["dataset_id"] = dataset_id
        df.rename(columns={"Table Name": "table_name", "Column Name": "field_name"}, inplace=True)
        spark_df = spark.createDataFrame(df)
        spark_df.write.format("delta").mode("overwrite").saveAsTable("pet.ctrl_dataset_input_fields_vj")
        logger.info("Input fields written successfully.")
    except Exception as e:
        logger.exception(f"Failed to write to pet.ctrl_dataset_input_fields_vj: {e}")
        raise
=============================================================================
# main()
=============================================================================
from datetime import datetime
import os
from openpyxl import load_workbook
import pandas as pd
import logging
from pyspark.sql import SparkSession, Row
from pyspark.sql.types import StructType, StructField, StringType, BooleanType, TimestampType, MapType

# from pet_loader.readers import download_to_tmp
# from pet_loader.validators import (
#     validate_required_sheets,
#     validate_specification_format,
#     check_database_exists,
#     check_table_and_column_exist,
#     validate_incremental_timestamp_field,
#     get_last_process_map
# )
# from pet_loader.writers import write_config, write_input_fields

logger = logging.getLogger("PETFileLoader")

class PETFileLoader:
    def __init__(self, spark: SparkSession):
        self.spark = spark

    def LoadConfigurationsForDataset(self, dataset_id: str, source_file_path: str, overwrite: bool = False) -> dict:
        results = {
            "dataset_id": dataset_id,
            "status": "FAILED",
            "errors": [],
            "config_written": False,
            "input_fields_written": False
        }
        try:
            # Step 1: Download & Load Excel Workbook
            local_path = download_to_tmp(source_file_path)
            wb = load_workbook(local_path, data_only=True)

            # Step 2: Validate sheets
            validate_required_sheets(wb.sheetnames)
            

            # Step 3: Specification format
            spec_df = pd.read_excel(local_path, sheet_name="Specification")
            spec_data = validate_specification_format(spec_df)

            if spec_data.get("ExternalID", "").upper() != dataset_id.upper():
                raise ValueError("Dataset ID mismatch between param and file")

            # Step 4: Parse params
            param_df = pd.read_excel(local_path, sheet_name="DPS-CDP Params")
            param_data = {str(k).strip(): str(v).strip() if pd.notnull(v) else None for k, v in zip(param_df.iloc[:, 0], param_df.iloc[:, 1]) if k}

            # Step 5: Skip if not enabled
            if param_data.get("Is Enabled", "TRUE").strip().upper() == "FALSE":
                raise Exception(f"Dataset {dataset_id} is disabled (Is Enabled = FALSE). Skipping load.")

            db_name = param_data.get("Database Name")
            if not db_name:
                raise ValueError("Database Name missing")
            if not check_database_exists(self.spark, db_name):
                raise ValueError(f"Database '{db_name}' not found")

            # Step 6: Validate incremental timestamp field
            validate_incremental_timestamp_field(self.spark, db_name, param_data.get("Incremental Timestamp Field"))

            # Step 7: Input fields
            input_fields_df = pd.read_excel(local_path, sheet_name="Input Fields")
            input_fields_df = input_fields_df.iloc[:, :2].dropna(how="all")
            input_fields_df.columns = [col.strip() for col in input_fields_df.columns]

            for _, row in input_fields_df.iterrows():
                table = str(row["Table Name"]).strip()
                column = str(row["Column Name"]).strip()
                if not check_table_and_column_exist(self.spark, db_name, table, column):
                    raise ValueError(f"Table or column not found: {db_name}.{table}.{column}")

            # Step 8: Compute last_process_datetime
            last_process_map = get_last_process_map(self.spark, dataset_id, param_data.get("Start Process From Date"))

            # Step 9: Write to config table
            config_written = write_config(self.spark, dataset_id, spec_data, param_data, db_name, last_process_map, overwrite)
            results["config_written"] = config_written

            # Step 10: Write to input fields
            input_fields_written = write_input_fields(self.spark, input_fields_df, dataset_id)
            results["input_fields_written"] = input_fields_written

            results["status"] = "SUCCESS"

        except Exception as e:
            logger.exception(f"Error during config load: {str(e)}")
            results["errors"].append(str(e))
        return results
