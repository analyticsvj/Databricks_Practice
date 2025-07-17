!pip install openpyxl==3.0.10

from datetime import datetime
import os
import re
import logging
import pandas as pd
from openpyxl import load_workbook
from pyspark.sql import SparkSession, Row
from pyspark.sql.types import *

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("pet_domain0_transformations_config")

class pet_domain0_transformations_config:
    def __init__(self, spark: SparkSession):
        self.spark = spark

    @staticmethod
    def validate_required_sheets(sheetnames):
        required = {"Specification", "Input Fields", "DPS-CDP Params"}
        missing = required - set(sheetnames)
        if missing:
            raise ValueError(f"Missing sheets: {missing}")

    @staticmethod
    def validate_specification_format(sheet):
        expected = [
            "Source Tenant", "Dataset Name", "Data Sharing Agreement (DSA)", "Purpose",
            "Privacy Domain - Encryption Key", "Destination Tenant", "ExternalID",
            "Allow Missing Table", "Priority", "Policy"
        ]
        rows = list(sheet.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True))
        if len(rows) < 10:
            raise ValueError(f"Specification must contain 10 rows, found {len(rows)}")
        for i, key in enumerate(expected):
            actual_key, actual_val = rows[i]
            if actual_key != key:
                raise ValueError(f"Row {i+1}: Expected '{key}', found '{actual_key}'")
            if actual_val is None or str(actual_val).strip() == "":
                raise ValueError(f"Row {i+1}: Missing value for '{key}'")

    @staticmethod
    def validate_input_fields(df):
        if df.columns.tolist()[:2] != ["Table Name", "Column Name"]:
            raise ValueError(f"Expected headers 'Table Name', 'Column Name', got {df.columns.tolist()[:2]}")
        if df.shape[0] < 1:
            raise ValueError("Input Fields must have more than one row")
        for i, row in df.iterrows():
            if not str(row["Column Name"]).strip():
                raise ValueError(f"Row {i+2}: Column Name is empty")

    def download_to_tmp(self, source_path):
        filename = os.path.basename(source_path)
        dbfs_path = f"dbfs:/tmp/{filename}"
        print(f"Copying to {dbfs_path}")
        dbutils.fs.cp(source_path, dbfs_path, recurse=False)
        return f"/dbfs/tmp/{filename}"

    def validate_dps_cdp_params(self, params, db_name):
        lt = params["load_type"].upper()
        sched = params.get("schedule")
        week = params.get("week_days")
        dates = params.get("dates_of_month")

        if lt not in {"INCREMENTAL", "FULL"}:
            raise ValueError("load_type must be 'INCREMENTAL' or 'FULL'")

        if lt == "INCREMENTAL":
            if not params.get("incremental_timestamp_field"):
                raise ValueError("incremental_timestamp_field required for INCREMENTAL load")
            if sched and sched.upper() == "DAILY":
                if not dates:
                    raise ValueError("dates_of_month required for DAILY")
                if week:
                    raise ValueError("week_days must be empty for DAILY")
            elif sched and sched.upper() == "WEEKLY":
                if not week:
                    raise ValueError("week_days required for WEEKLY")
                if dates:
                    raise ValueError("dates_of_month must be empty for WEEKLY")
        else:
            if any([params.get(k) for k in [
                "incremental_timestamp_field", "incremental_header_to_tables_link_field", "schedule", "week_days", "dates_of_month"]]):
                raise ValueError("All incremental-related fields must be empty for FULL load")

        if not params.get("pet_dataset_id"):
            raise ValueError("pet_dataset_id is required")

        method = params.get("incremental_timestamp_method", "")
        if method.upper() not in {"TIMESTAMP_HEADER_TABLE", "TIMESTAMP_DATA_TABLE"}:
            raise ValueError("incremental_timestamp_method must be valid")

        if method.lower() == "timestamp_header_table":
            val = params.get("incremental_header_to_tables_link_field")
            if not val or "." not in val:
                raise ValueError("Link field must be in <table>.<field> format")
            tbl, fld = val.split(".", 1)
            if not self.spark.sql(f"SHOW TABLES IN {db_name}").filter(f"tableName = '{tbl}'").count():
                raise ValueError(f"Table '{tbl}' not found in database '{db_name}'")
            fields = [f.name for f in self.spark.table(f"{db_name}.{tbl}").schema.fields]
            if fld not in fields:
                raise ValueError(f"Field '{fld}' not in table '{tbl}'")

        if sched and sched.capitalize() not in {"Daily", "Weekly", "Monthly"}:
            raise ValueError("Invalid schedule")

        if week:
            valid = {"Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"}
            days = {d.strip().capitalize() for d in week.split(",")}
            if not days.issubset(valid):
                raise ValueError(f"Invalid week_days: {days - valid}")

        if dates:
            for d in dates.split(","):
                if not d.strip().isdigit() or not 1 <= int(d.strip()) <= 28:
                    raise ValueError(f"Invalid date: {d.strip()}")

    def LoadConfigurationsForDataset(self, dataset_id, source_file_path, overwrite=False):
        print("Starting configuration load...")
        local_path = self.download_to_tmp(source_file_path)
        try:
            wb = load_workbook(local_path, data_only=True)
            self.validate_required_sheets(wb.sheetnames)
            self.validate_specification_format(wb["Specification"])

            spec = wb["Specification"]
            dataset_name = str(spec["B2"].value).strip()
            dataset_id_file = str(spec["B7"].value).strip().upper()
            if dataset_id_file != dataset_id.upper():
                raise ValueError(f"Mismatch: parameter '{dataset_id}' vs file '{dataset_id_file}'")

            params = wb["DPS-CDP Params"]
            param_data = {
                key.strip().lower(): str(value).strip() if value is not None else None
                for key, value in (row for row in params.iter_rows(min_row=2, max_row=15, max_col=2, values_only=True))
                if key
            }

            db_name = param_data.get("database_name")
            if not db_name:
                raise ValueError("database_name is missing")

            if db_name.lower() not in [db.databaseName.lower() for db in self.spark.sql("SHOW DATABASES").collect()]:
                raise ValueError(f"Hive database '{db_name}' not found")

            self.validate_dps_cdp_params(param_data, db_name)

            df = pd.read_excel(local_path, sheet_name="Input Fields", engine="openpyxl").iloc[:, :2].dropna(how="all")
            df.columns = [c.strip() for c in df.columns]
            self.validate_input_fields(df)

            for _, row in df.iterrows():
                tbl = row["Table Name"].strip()
                coln = row["Column Name"].strip()
                full_tbl = f"{db_name}.{tbl}"
                if not self.spark.sql(f"SHOW TABLES IN {db_name}").filter(f"tableName = '{tbl}'").count():
                    raise ValueError(f"Table '{tbl}' not found")
                schema = [f.name for f in self.spark.table(full_tbl).schema.fields]
                if coln not in schema:
                    raise ValueError(f"Column '{coln}' not found in table '{tbl}'")

            if self.spark.sql(f"SELECT * FROM PET.CTRL_dataset_config WHERE dataset_id = '{dataset_id}'").count():
                if not overwrite:
                    raise Exception(f"Dataset {dataset_id} exists. Use overwrite=True")
                self.spark.sql(f"DELETE FROM PET.CTRL_dataset_config WHERE dataset_id = '{dataset_id}'")
                self.spark.sql(f"DELETE FROM PET.CTRL_dataset_input_fields WHERE dataset_id = '{dataset_id}'")

            now = datetime.now()
            row_obj = Row(
                dataset_id=dataset_id,
                dataset_name=dataset_name,
                source_database_name=param_data["database_name"],
                load_type=param_data["load_type"],
                is_enabled=True,
                pet_dataset_id=param_data["pet_dataset_id"],
                incremental_timestamp_method=param_data["incremental_timestamp_method"],
                incremental_timestamp_field=param_data["incremental_timestamp_field"],
                incremental_header_to_tables_link_field=param_data.get("incremental_header_to_tables_link_field"),
                schedule=param_data.get("schedule"),
                week_days=param_data.get("week_days"),
                dates_of_month=param_data.get("dates_of_month"),
                dataset_owner=param_data["dataset_owner"],
                modified_date=now,
                dataset_modified_by_user=dbutils.notebook.entry_point.getDbutils().notebook().getContext().tags().apply('user'),
                comments=param_data.get("comments")
            )

            schema = StructType([
                StructField("dataset_id", StringType(), False),
                StructField("dataset_name", StringType(), False),
                StructField("source_database_name", StringType(), False),
                StructField("load_type", StringType(), False),
                StructField("is_enabled", BooleanType(), False),
                StructField("pet_dataset_id", StringType(), False),
                StructField("incremental_timestamp_method", StringType(), True),
                StructField("incremental_timestamp_field", StringType(), True),
                StructField("incremental_header_to_tables_link_field", StringType(), True),
                StructField("schedule", StringType(), True),
                StructField("week_days", StringType(), True),
                StructField("dates_of_month", StringType(), True),
                StructField("dataset_owner", StringType(), False),
                StructField("modified_date", TimestampType(), False),
                StructField("dataset_modified_by_user", StringType(), False),
                StructField("comments", StringType(), True)
            ])

            self.spark.createDataFrame([row_obj], schema).write.mode("append").saveAsTable("PET.CTRL_dataset_config")

            input_rows = [
                Row(dataset_id=dataset_id, table_name=r["Table Name"].strip(), field_name=r["Column Name"].strip())
                for _, r in df.iterrows()
            ]

            input_schema = StructType([
                StructField("dataset_id", StringType(), False),
                StructField("table_name", StringType(), False),
                StructField("field_name", StringType(), False)
            ])

            self.spark.createDataFrame(input_rows, input_schema).coalesce(1).write.mode("append").saveAsTable("PET.CTRL_dataset_input_fields")
            print(f"✅ Dataset '{dataset_id}' onboarded successfully")

        finally:
            if os.path.exists(local_path):
                try:
                    os.remove(local_path)
                except Exception as e:
                    print(f"Warning: cleanup failed: {e}")
