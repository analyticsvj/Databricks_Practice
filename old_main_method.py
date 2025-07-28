from datetime import datetime
import os
from openpyxl import load_workbook
import pandas as pd
import logging
from pyspark.sql import SparkSession, Row
from pyspark.sql.types import StructType, StructField, StringType, BooleanType, TimestampType, MapType

# from pet_loader.readers import download_to_tmp
# from pet_loader.validators import (
#     validate_required_sheets,
#     validate_specification_format,
#     check_database_exists,
#     check_table_and_column_exist,
#     validate_incremental_timestamp_field,
#     get_last_process_map
# )
# from pet_loader.writers import write_config, write_input_fields


logger = logging.getLogger("PETFileLoader")

class PETFileLoader:
    def __init__(self, spark: SparkSession):
        self.spark = spark

    def LoadConfigurationsForDataset(self, dataset_id: str, source_file_path: str, overwrite: bool = False) -> dict:
        results = {
            "dataset_id": dataset_id,
            "status": "FAILED",
            "errors": [],
            "config_written": False,
            "input_fields_written": False
        }
        try:
            # # Validate file existence
            # if not os.path.exists(source_file_path):
            #     raise FileNotFoundError(f"The file was not found at: {source_file_path}")

            # Validate file extension
            if not source_file_path.lower().endswith(".xlsx"):
                raise ValueError(f"Invalid file type for {source_file_path}. Only .xlsx files are supported.")

            # Step 1: Download & Load Excel Workbook
            local_path = download_to_tmp(source_file_path)
            wb = load_workbook(local_path, data_only=True)

            # Step 2: Validate sheets
            validate_required_sheets(wb.sheetnames)
            

            # Step 3: Specification format
            spec_data = validate_specification_format(wb["Specification"])

            if spec_data.get("ExternalID", "").upper() != dataset_id.upper():
                raise ValueError("Dataset ID mismatch between param and file")

            # Step 4: Parse params
            param_df = pd.read_excel(local_path, sheet_name="DPS-CDP Params")
            param_data = {str(k).strip(): str(v).strip() if pd.notnull(v) else None for k, v in zip(param_df.iloc[:, 0], param_df.iloc[:, 1]) if k}

            # Step 5: Skip if not enabled
            if param_data.get("Is Enabled", "TRUE").strip().upper() == "FALSE":
                raise Exception(f"Dataset {dataset_id} is disabled (Is Enabled = FALSE). Skipping load.")
            
            # If overwrite is False, skip
            # if not overwrite:
            #     logger.warning(f"Dataset '{dataset_id}' is already configured and overwrite is set to False. Skipping load.")
            #     return {"dataset_id": dataset_id, "status": "SKIPPED", "reason": "Overwrite disabled"}
            if not overwrite:
                existing_datasets = self.spark.sql("SELECT DISTINCT dataset_id FROM pet.ctrl_dataset_config")
                if dataset_id in [row['dataset_id'] for row in existing_datasets.collect()]:
                    logger.warning(f"Dataset '{dataset_id}' already exists and overwrite is set to False. Skipping load.")
                    return {"dataset_id": dataset_id, "status": "SKIPPED", "reason": "Overwrite disabled"}



            db_name = param_data.get("Database Name")
            if not db_name:
                raise ValueError("Database Name missing")
            if not check_database_exists(self.spark, db_name):
                raise ValueError(f"Database '{db_name}' not found")

            # # Step 6: Validate incremental timestamp field
            validate_incremental_timestamp_field(self.spark, db_name, param_data.get("Incremental Timestamp Field"))



            # Step 7: Input fields
            input_fields_df = pd.read_excel(local_path, sheet_name="Input Fields")
            input_fields_df = input_fields_df.iloc[:, :2].dropna(how="all")
            input_fields_df.columns = [col.strip() for col in input_fields_df.columns]

            for _, row in input_fields_df.iterrows():
                table = str(row["Table Name"]).strip()
                column = str(row["Column Name"]).strip()
                if not check_table_and_column_exist(self.spark, db_name, table, column):
                    raise ValueError(f"Table or column not found: {db_name}.{table}.{column}")

            # Step 8: Compute last_process_datetime
            last_process_map = get_last_process_map(self.spark, dataset_id, param_data.get("Start Process From Date"))

            # Step 9: Write to config table
            config_written = write_config(self.spark, dataset_id, spec_data, param_data, db_name, last_process_map, overwrite)
            results["config_written"] = config_written

            # Step 10: Write to input fields
            input_fields_written = write_input_fields_table(self.spark, input_fields_df, dataset_id)
            results["input_fields_written"] = input_fields_written

            results["status"] = "SUCCESS"

        except Exception as e:
            logger.exception(f"Error during config load: {str(e)}")
            results["errors"].append(str(e))
        return results
