import pytest
from unittest.mock import MagicMock, patch
from pyspark.sql import SparkSession
from pyspark.sql.types import StructType, StringType, BooleanType, TimestampType
from pyspark.sql import Row
from datetime import datetime
from your_module import PetDomain0TransformationsConfig  # replace 'your_module' with actual module name

# Fixture to create a SparkSession for tests
@pytest.fixture(scope="module")
def spark():
    return SparkSession.builder.master("local[1]").appName("pytest").getOrCreate()

@pytest.fixture
def config_loader(spark):
    return PetDomain0TransformationsConfig(spark)

def test_validate_required_sheets_success():
    sheets = ["Specification", "Input Fields", "DPS-CDP Params", "Extra"]
    assert PetDomain0TransformationsConfig.validate_required_sheets(sheets) is True

def test_validate_required_sheets_missing():
    sheets = ["Specification", "Input Fields"]
    with pytest.raises(ValueError, match="Missing sheets"):
        PetDomain0TransformationsConfig.validate_required_sheets(sheets)

def test_validate_specification_format_success(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Specification"
    keys = [
        "Source Tenant", "Dataset Name", "Data Sharing Agreement (DSA)", "Purpose",
        "Privacy Domain - Encryption Key", "Destination Tenant", "ExternalID",
        "Allow Missing Table", "Priority", "Policy"
    ]
    for i, key in enumerate(keys, 1):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value="some_value")

    file_path = tmp_path / "spec.xlsx"
    wb.save(file_path)

    wb_loaded = load_workbook(filename=file_path)
    sheet = wb_loaded["Specification"]
    assert PetDomain0TransformationsConfig.validate_specification_format(sheet) is True

def test_validate_specification_format_fail_missing_value(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Specification"
    keys = [
        "Source Tenant", "Dataset Name", "Data Sharing Agreement (DSA)", "Purpose",
        "Privacy Domain - Encryption Key", "Destination Tenant", "ExternalID",
        "Allow Missing Table", "Priority", "Policy"
    ]
    for i, key in enumerate(keys, 1):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=None if i==3 else "some_value")

    file_path = tmp_path / "spec_fail.xlsx"
    wb.save(file_path)

    wb_loaded = load_workbook(filename=file_path)
    sheet = wb_loaded["Specification"]

    with pytest.raises(ValueError, match="Missing value for 'Data Sharing Agreement"):
        PetDomain0TransformationsConfig.validate_specification_format(sheet)

def test_validate_input_fields_success(config_loader):
    import pandas as pd
    df = pd.DataFrame({
        "Table Name": ["table1", "table2"],
        "Column Name": ["col1", "col2"],
        "Other": [1, 2]
    })
    assert config_loader.validate_input_fields(df) is True

def test_validate_input_fields_fail_wrong_headers(config_loader):
    import pandas as pd
    df = pd.DataFrame({
        "Wrong Name": ["table1"],
        "Column Name": ["col1"]
    })
    with pytest.raises(ValueError, match="Expected headers"):
        config_loader.validate_input_fields(df)

def test_validate_input_fields_fail_empty_column_name(config_loader):
    import pandas as pd
    df = pd.DataFrame({
        "Table Name": ["table1"],
        "Column Name": [""]
    })
    with pytest.raises(ValueError, match="Column Name is empty"):
        config_loader.validate_input_fields(df)

def test_validate_dps_cdp_params_success(config_loader, spark):
    params = {
        "load_type": "INCREMENTAL",
        "schedule": "Daily",
        "week_days": "",
        "dates_of_month": "1,15",
        "incremental_timestamp_field": "ts",
        "pet_dataset_id": "pet123",
        "incremental_timestamp_method": "TIMESTAMP_HEADER_TABLE",
        "incremental_header_to_tables_link_field": "table1.field1"
    }

    # Mock the table existence and fields
    def sql_mock(query):
        class Result:
            def count(self_inner):
                # simulate table found
                return 1
            def filter(self_inner, cond):
                return self_inner
        return Result()

    def table_mock(name):
        class Tbl:
            schema = StructType([
                StructField("field1", StringType()),
                StructField("field2", StringType()),
            ])
        return Tbl()

    config_loader.spark.sql = MagicMock(side_effect=sql_mock)
    config_loader.spark.table = MagicMock(side_effect=table_mock)

    assert config_loader.validate_dps_cdp_params(params, "db_name") is True

def test_validate_dps_cdp_params_fail_invalid_load_type(config_loader):
    params = {"load_type": "UNKNOWN"}
    with pytest.raises(ValueError, match="load_type must be 'INCREMENTAL' or 'FULL'"):
        config_loader.validate_dps_cdp_params(params, "db")

def test_remove_existing_dataset_success(config_loader):
    config_loader.spark.sql = MagicMock(side_effect=lambda q: MagicMock(count=MagicMock(return_value=1)))
    config_loader.spark.sql = MagicMock()
    config_loader.spark.sql.return_value.count.return_value = 1
    # Should return True as dataset exists and deleted
    assert config_loader.remove_existing_dataset("dataset1") is True

def test_remove_existing_dataset_fail(config_loader):
    config_loader.spark.sql = MagicMock(side_effect=Exception("DB error"))
    assert config_loader.remove_existing_dataset("dataset1") is False

def test_write_dataset_config_success(config_loader):
    param_data = {
        "database_name": "db",
        "load_type": "FULL",
        "pet_dataset_id": "pet1",
        "incremental_timestamp_method": "method",
        "incremental_timestamp_field": "field",
        "incremental_header_to_tables_link_field": "link",
        "schedule": "Daily",
        "week_days": "Monday",
        "dates_of_month": "1",
        "dataset_owner": "owner",
        "comments": "some comments"
    }
    config_loader.spark.createDataFrame = MagicMock(return_value=MagicMock(write=MagicMock(mode=MagicMock(return_value=MagicMock(saveAsTable=MagicMock())))))

    # Mock dbutils context for user
    with patch("builtins.dbutils", create=True) as mock_dbutils:
        mock_dbutils.notebook.entry_point.getDbutils().notebook().getContext().tags().apply.return_value = "user1"
        result = config_loader.write_dataset_config("id1", "dataset_name", param_data)
    assert result is True

def test_write_dataset_config_fail(config_loader):
    param_data = {}
    config_loader.spark.createDataFrame = MagicMock(side_effect=Exception("Failed to create DF"))
    result = config_loader.write_dataset_config("id1", "dataset_name", param_data)
    assert result is False

@patch("your_module.PetDomain0TransformationsConfig.validate_required_sheets")
@patch("your_module.PetDomain0TransformationsConfig.validate_specification_format")
@patch("your_module.PetDomain0TransformationsConfig.validate_input_fields")
@patch("your_module.PetDomain0TransformationsConfig.validate_dps_cdp_params")
@patch("your_module.PetDomain0TransformationsConfig.remove_existing_dataset")
@patch("your_module.PetDomain0TransformationsConfig.write_dataset_config")
def test_LoadConfigurationsForDataset_success(
    mock_write, mock_remove, mock_dps_cdp, mock_input, mock_spec, mock_req, config_loader, tmp_path
):
    mock_req.return_value = True
    mock_spec.return_value = True
    mock_input.return_value = True
    mock_dps_cdp.return_value = True
    mock_remove.return_value = True
    mock_write.return_value = True

    # You may want to mock actual reading of the file or use tmp_path to create a dummy file

    result = config_loader.LoadConfigurationsForDataset(
        dataset_id="ds1",
        source_file_path=str(tmp_path / "dummy.xlsx"),
        overwrite=True
    )
    assert result is True

def test_LoadConfigurationsForDataset_fail_on_validation(config_loader):
    with patch.object(config_loader, "validate_required_sheets", side_effect=ValueError("Missing sheets")):
        result = config_loader.LoadConfigurationsForDataset(
            dataset_id="ds1",
            source_file_path="fake.xlsx",
            overwrite=True
        )
        assert result is False
