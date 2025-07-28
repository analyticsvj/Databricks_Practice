from datetime import datetime
import os
from openpyxl import load_workbook
import pandas as pd
import logging
from pyspark.sql import SparkSession, Row
from pyspark.sql.types import StructType, StructField, StringType, BooleanType, TimestampType, MapType

from readers import download_to_tmp
from validators import (
    validate_required_sheets,
    validate_specification_format,
    validate_external_id,
    validate_dataset_enabled,
    validate_database_name,
    validate_incremental_timestamp_field,
    validate_input_fields,
    get_last_process_map
)
from writers import write_config, write_input_fields

logger = logging.getLogger("PETFileLoader")

class PETFileLoader:
    def __init__(self, spark: SparkSession):
        self.spark = spark

    def LoadConfigurationsForDataset(self, dataset_id: str, source_file_path: str, overwrite: bool = False) -> dict:
        results = {
            "dataset_id": dataset_id,
            "status": "FAILED",
            "errors": [],
            "config_written": False,
            "input_fields_written": False
        }
        try:
            # Step 1: Download & Load Excel Workbook
            local_path = download_to_tmp(source_file_path)
            wb = load_workbook(local_path, data_only=True)

            # Step 2: Validate sheets
            validate_required_sheets(wb.sheetnames)
            

            # Step 3: Specification format
            spec_data = validate_specification_format(wb["Specification"])

            # Step 4: Validate External ID
            validate_external_id(spec_data, dataset_id)
            
            # Step 5: Validate Dataset Name
            validate_dataset_name(spec_data, dataset_id)

            # Step 6: Parse params
            param_df = pd.read_excel(local_path, sheet_name="DPS-CDP Params")
            param_data = {str(k).strip(): str(v).strip() if pd.notnull(v) else None for k, v in zip(param_df.iloc[:, 0], param_df.iloc[:, 1]) if k}

            # Step 7: Check if dataset is enabled
            validate_dataset_enabled(param_data, dataset_id)

            # Step 8: Validate database
            db_name = validate_database_name(self.spark, param_data)

            # Step 9: Validate DPS-CDP parameters
            validate_dps_cdp_params(self.spark, param_data, db_name)

            # Step 10: Validate incremental timestamp field
            validate_incremental_timestamp_field(self.spark, db_name, param_data.get("Incremental Timestamp Field"))

            # Step 11: Validate input fields
            input_fields_df = pd.read_excel(local_path, sheet_name="Input Fields")
            validate_input_fields(self.spark, input_fields_df, db_name)

            # Step 12: Compute last_process_datetime
            last_process_map = get_last_process_map(self.spark, dataset_id, param_data.get("Start Process From Date"))

            # Step 13: Write to config table
            config_written = write_config(self.spark, dataset_id, spec_data, param_data, db_name, last_process_map, overwrite)
            results["config_written"] = config_written

            # Step 14: Write to input fields
            input_fields_written = write_input_fields(self.spark, input_fields_df, dataset_id)
            results["input_fields_written"] = input_fields_written

            results["status"] = "SUCCESS"

        except Exception as e:
            logger.exception(f"Error during config load: {str(e)}")
            results["errors"].append(str(e))
        return results
