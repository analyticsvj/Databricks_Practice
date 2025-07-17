from datetime import datetime
import os
import re
import logging
import pandas as pd
from openpyxl import load_workbook
from pyspark.sql import SparkSession, Row
from pyspark.sql.types import *

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("pet_domain0_transformations_config")

class PetDomain0TransformationsConfig:
    def __init__(self, spark: SparkSession):
        self.spark = spark

    def LoadConfigurationsForDataset(self, dataset_id: str, source_file_path: str, overwrite: bool):
        try:
            local_path = self.download_to_tmp(source_file_path)
            wb = load_workbook(local_path, data_only=True)
            sheetnames = wb.sheetnames
            self.validate_required_sheets(sheetnames)

            spec_sheet = wb["Specification"]
            self.validate_specification_format(spec_sheet)

            df_inputs = pd.read_excel(local_path, sheet_name="Input Fields")
            self.validate_input_fields(df_inputs)

            df_params = pd.read_excel(local_path, sheet_name="DPS-CDP Params")
            params_dict = dict(zip(df_params.iloc[:, 0], df_params.iloc[:, 1]))
            db_name = params_dict.get("database_name")
            self.validate_dps_cdp_params(params_dict, db_name)

            if overwrite:
                removed = self.remove_existing_dataset(dataset_id)
                if not removed:
                    logger.warning("No existing dataset found to remove.")

            dataset_name = spec_sheet["B2"].value
            if not dataset_name:
                raise ValueError("Dataset Name is missing in Specification sheet")  # missed validation

            config_written = self.write_dataset_config(dataset_id, dataset_name, params_dict)
            if not config_written:
                raise Exception("Failed to write dataset configuration")

            logger.info("Configuration successfully loaded for dataset: %s", dataset_id)
            return True
        except Exception as e:
            logger.error(f"Failed to load configuration for dataset {dataset_id}: {e}")
            return False

    @staticmethod
    def validate_required_sheets(sheetnames):
        required = {"Specification", "Input Fields", "DPS-CDP Params"}
        missing = required - set(sheetnames)
        if missing:
            raise ValueError(f"Missing sheets: {missing}")
        return True

    @staticmethod
    def validate_specification_format(sheet):
        expected = [
            "Source Tenant", "Dataset Name", "Data Sharing Agreement (DSA)", "Purpose",
            "Privacy Domain - Encryption Key", "Destination Tenant", "ExternalID",
            "Allow Missing Table", "Priority", "Policy"
        ]
        rows = list(sheet.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True))
        if len(rows) < 10:
            raise ValueError(f"Specification must contain 10 rows, found {len(rows)}")
        for i, key in enumerate(expected):
            actual_key, actual_val = rows[i]
            if actual_key != key:
                raise ValueError(f"Row {i+1}: Expected '{key}', found '{actual_key}'")
            if actual_val is None or str(actual_val).strip() == "":
                raise ValueError(f"Row {i+1}: Missing value for '{key}'")
        return True

    @staticmethod
    def validate_input_fields(df):
        if df.columns.tolist()[:2] != ["Table Name", "Column Name"]:
            raise ValueError(f"Expected headers 'Table Name', 'Column Name', got {df.columns.tolist()[:2]}")
        if df.shape[0] < 1:
            raise ValueError("Input Fields must have more than one row")
        for i, row in df.iterrows():
            if not str(row["Column Name"]).strip():
                raise ValueError(f"Row {i+2}: Column Name is empty")
        return True

    def download_to_tmp(self, source_path):
        try:
            filename = os.path.basename(source_path)
            dbfs_path = f"dbfs:/tmp/{filename}"
            dbutils.fs.cp(source_path, dbfs_path, recurse=False)
            return f"/dbfs/tmp/{filename}"
        except Exception as e:
            logger.error(f"Failed to download file: {e}")
            raise

    def validate_dps_cdp_params(self, params, db_name):
        try:
            lt = params["load_type"].upper()
            sched = params.get("schedule")
            week = params.get("week_days")
            dates = params.get("dates_of_month")

            if lt not in {"INCREMENTAL", "FULL"}:
                raise ValueError("load_type must be 'INCREMENTAL' or 'FULL'")

            if lt == "INCREMENTAL":
                if not params.get("incremental_timestamp_field"):
                    raise ValueError("incremental_timestamp_field required for INCREMENTAL load")

                if sched and sched.upper() == "DAILY":
                    if not dates:
                        raise ValueError("dates_of_month required for DAILY")
                    if week:
                        raise ValueError("week_days must be empty for DAILY")
                elif sched and sched.upper() == "WEEKLY":
                    if not week:
                        raise ValueError("week_days required for WEEKLY")
                    if dates:
                        raise ValueError("dates_of_month must be empty for WEEKLY")
            else:
                if any([params.get(k) for k in [
                    "incremental_timestamp_field", "incremental_header_to_tables_link_field",
                    "schedule", "week_days", "dates_of_month"]]):
                    raise ValueError("All incremental-related fields must be empty for FULL load")

            if not params.get("pet_dataset_id"):
                raise ValueError("pet_dataset_id is required")

            method = params.get("incremental_timestamp_method", "")
            if method.upper() not in {"TIMESTAMP_HEADER_TABLE", "TIMESTAMP_DATA_TABLE"}:
                raise ValueError("incremental_timestamp_method must be valid")

            if method.lower() == "timestamp_header_table":
                val = params.get("incremental_header_to_tables_link_field")
                if not val or "." not in val:
                    raise ValueError("Link field must be in <table>.<field> format")
                tbl, fld = val.split(".", 1)
                if not self.spark.sql(f"SHOW TABLES IN {db_name}").filter(f"tableName = '{tbl}'").count():
                    raise ValueError(f"Table '{tbl}' not found in database '{db_name}'")
                fields = [f.name for f in self.spark.table(f"{db_name}.{tbl}").schema.fields]
                if fld not in fields:
                    raise ValueError(f"Field '{fld}' not in table '{tbl}'")

            if sched and sched.capitalize() not in {"Daily", "Weekly", "Monthly"}:
                raise ValueError("Invalid schedule")

            if week:
                valid = {"Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"}
                days = {d.strip().capitalize() for d in week.split(",")}
                if not days.issubset(valid):
                    raise ValueError(f"Invalid week_days: {days - valid}")

            if dates:
                for d in dates.split(","):
                    if not d.strip().isdigit() or not 1 <= int(d.strip()) <= 28:
                        raise ValueError(f"Invalid date: {d.strip()}")

            return True
        except Exception as e:
            logger.error(f"Validation failed: {e}")
            raise

    def remove_existing_dataset(self, dataset_id):
        try:
            removed = False
            if self.spark.sql(f"SELECT * FROM PET.CTRL_dataset_config WHERE dataset_id = '{dataset_id}'").count():
                self.spark.sql(f"DELETE FROM PET.CTRL_dataset_config WHERE dataset_id = '{dataset_id}'")
                self.spark.sql(f"DELETE FROM PET.CTRL_dataset_input_fields WHERE dataset_id = '{dataset_id}'")
                removed = True
            return removed
        except Exception as e:
            logger.error(f"Failed to remove existing dataset: {e}")
            return False

    def write_dataset_config(self, dataset_id, dataset_name, param_data):
        try:
            now = datetime.now()
            row_obj = Row(
                dataset_id=dataset_id,
                dataset_name=dataset_name,
                source_database_name=param_data["database_name"],
                load_type=param_data["load_type"],
                is_enabled=True,
                pet_dataset_id=param_data["pet_dataset_id"],
                incremental_timestamp_method=param_data["incremental_timestamp_method"],
                incremental_timestamp_field=param_data["incremental_timestamp_field"],
                incremental_header_to_tables_link_field=param_data.get("incremental_header_to_tables_link_field"),
                schedule=param_data.get("schedule"),
                week_days=param_data.get("week_days"),
                dates_of_month=param_data.get("dates_of_month"),
                dataset_owner=param_data["dataset_owner"],
                modified_date=now,
                dataset_modified_by_user=dbutils.notebook.entry_point.getDbutils().notebook().getContext().tags().apply('user'),
                comments=param_data.get("comments")
            )

            schema = StructType([
                StructField("dataset_id", StringType(), False),
                StructField("dataset_name", StringType(), False),
                StructField("source_database_name", StringType(), False),
                StructField("load_type", StringType(), False),
                StructField("is_enabled", BooleanType(), False),
                StructField("pet_dataset_id", StringType(), False),
                StructField("incremental_timestamp_method", StringType(), True),
                StructField("incremental_timestamp_field", StringType(), True),
                StructField("incremental_header_to_tables_link_field", StringType(), True),
                StructField("schedule", StringType(), True),
                StructField("week_days", StringType(), True),
                StructField("dates_of_month", StringType(), True),
                StructField("dataset_owner", StringType(), False),
                StructField("modified_date", TimestampType(), False),
                StructField("dataset_modified_by_user", StringType(), False),
                StructField("comments", StringType(), True)
            ])

            df = self.spark.createDataFrame([row_obj], schema)
            df.write.mode("append").saveAsTable("PET.CTRL_dataset_config")
            return True
        except Exception as e:
            logger.error(f"Failed to write dataset config: {e}")
            return False
